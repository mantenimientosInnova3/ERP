
# compras/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .forms import ProductoForm
from .models import OrdenCompra, DetalleOrden, Producto
from .forms import OrdenCompraForm, DetalleOrdenForm
from django.http import HttpResponse
from django.db.models import Sum, Q, F, DecimalField, ExpressionWrapper
from .models import CentroCosto
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.template.loader import get_template
from .models import OrdenCompra, DetalleOrden
from decimal import Decimal
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import RequisicionForm, DetalleRequisicionForm
from .models import Requisicion, DetalleRequisicion, OrdenCompra, DetalleOrden
from django.contrib import messages
from django.utils import timezone
from .models import MovimientoInventario
from django.conf import settings
import os
import base64

def es_compras(user):
    return user.is_staff

def es_admin(user):
    return user.is_superuser  # opcional, si usas superusuario como admin

@login_required
def crear_requisicion(request):
    if request.method == "POST":
        req_form = RequisicionForm(request.POST)
        det_form = DetalleRequisicionForm(request.POST)
        if req_form.is_valid() and det_form.is_valid():
            requisicion = req_form.save(commit=False)
            requisicion.usuario = request.user
            requisicion.save()
            requisicion.consecutivo = f"{request.user.username.upper()}-{requisicion.id:03d}"
            requisicion.save()

            detalle = det_form.save(commit=False)
            detalle.requisicion = requisicion
            detalle.save()

            return redirect('mis_requisiciones')
    else:
        req_form = RequisicionForm()
        det_form = DetalleRequisicionForm()

    return render(request, 'compras/crear_requisicion.html', {
        'req_form': req_form,
        'det_form': det_form,
    })



@login_required
def home(request):
    usuario = request.user
    es_admin = usuario.is_superuser
    es_compras = usuario.groups.filter(name='Compras').exists()
    total_productos = Producto.objects.count()
    total_productos = Producto.objects.count()
    total_ordenes = OrdenCompra.objects.count()
    ultimas_ordenes = OrdenCompra.objects.order_by('-fecha')[:5]
    
    return render(request, 'compras/home.html', {
        'usuario': request.user,
        'es_admin': es_admin,
        'es_compras': es_compras,
        'total_productos': total_productos,
        'total_ordenes': total_ordenes,
        'ultimas_ordenes': ultimas_ordenes,
    })


@login_required
def mis_requisiciones(request):
    requisiciones = Requisicion.objects.filter(usuario=request.user).order_by('-fecha')
    return render(request, 'compras/mis_requisiciones.html', {'requisiciones': requisiciones})

@login_required
@user_passes_test(es_compras)
def generar_orden_de_requisicion(request, id):
    requisicion = get_object_or_404(Requisicion, id=id, estado="APROBADA")
    detalles = requisicion.detalles.all()

    if request.method == "POST":
        form = OrdenCompraForm(request.POST)
        if form.is_valid():
            orden = form.save()

            for partida in detalles:
                DetalleOrden.objects.create(
                    orden=orden,
                    producto=partida.producto,       # ahora es FK a Producto
                    unidad=partida.unidad,
                    cantidad=partida.cantidad,
                    observaciones=partida.observaciones,
                    precio_unitario=0
                )
                
                # Esto es para actualizar inventario al momento de generar la orden
                producto =detalle.producto
                producto.cantidad += detalle.cantidad
                producto.save()

            messages.success(request, "Orden de compra generada con éxito.")
            return redirect('lista_ordenes')
    else:
        form = OrdenCompraForm()

    return render(request, 'compras/generar_orden_de_requisicion.html', {
        'requisicion': requisicion,
        'detalles': detalles,
        'form': form,
    })

@login_required
@user_passes_test(es_compras)
def exportar_reporte_pdf(request):
    
    logo_path = os.path.join(settings.STATIC_ROOT, 'logo.jpg')
    if os.path.exists(logo_path):
        logo_url = os.path.join(settings.STATIC_URL, 'logo.jpg')
    else:
        logo_url = None
    
    context = {
        'logo_url': logo_url,
    }
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    cc_id = request.GET.get('centro_costo')
    
    qs_ordenes = OrdenCompra.objects.all()
    if fecha_inicio:
        qs_ordenes = qs_ordenes.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs_ordenes = qs_ordenes.filter(fecha__lte=fecha_fin)
    if cc_id:
        qs_ordenes = qs_ordenes.filter(centro_costo__id=cc_id)
        
    cc_data = []
    for cc in CentroCosto.objects.all():
        if cc_id and str(cc.id) != cc_id:
            continue
        
        órdenes = OrdenCompra.objects.filter(centro_costo=cc)
        total_gasto = 0
        detalles = DetalleOrden.objects.filter(orden__in=órdenes)
        desglose = []
        for det in detalles:
            costo = det.cantidad * det.producto.precio
            total_gasto += costo
            desglose.append({
                'producto': det.producto.nombre,
                'cantidad': det.cantidad,
                'precio': det.producto.precio,
                'gasto': costo,
            })
        cc_data.append({
            'centro_costo': cc,
            'total_gasto': total_gasto,
            'detalles': desglose
        })
    total_general = sum(cc['total_gasto'] for cc in cc_data)
    
    
    print(f"Ruta del logo: {logo_path}")
    print(f"¿Existe el archivo?: {os.path.exists(logo_path)}")
     
    template = get_template('compras/reporte_pdf.html')    
    html = template.render({
        'cc_data': cc_data,
        'total_general': total_general,
        'now': timezone.now(),
        'logo_url': logo_url
    }, request)
    
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_gastos.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
        
    if pisa_status.err:
        return HttpResponse('Error generando PDF')
    return response


@login_required
@user_passes_test(es_compras)
def exportar_reporte_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    cc_id = request.GET.get('centro_costo')
    
    qs_ordenes = OrdenCompra.objects.all()
    if fecha_inicio:
        qs_ordenes = qs_ordenes.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs_ordenes = qs_ordenes.filter(fecha__lte=fecha_fin)
    if cc_id:
        qs_ordenes = qs_ordenes.filter(centro_costo__id=cc_id)

    wb = Workbook()

    # Estilos
    header_fill = PatternFill("solid", fgColor="003366")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    moneda = NamedStyle(name="moneda")
    moneda.number_format = '"$"#,##0.00'
    if "moneda" not in wb.named_styles:
        wb.add_named_style(moneda)

    # ---------- HOJA 1: RESUMEN ----------
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    # Título resumen
    ws_resumen.merge_cells('A1:C1')
    ws_resumen['A1'] = "Resumen de gastos por centro de costo"
    ws_resumen['A1'].font = Font(size=14, bold=True)
    ws_resumen['A1'].alignment = Alignment(horizontal="center")

    # Encabezados resumen
    headers_res = ['Centro de Costo', 'Total gastado']
    for col, header in enumerate(headers_res, start=1):
        cell = ws_resumen.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_res = 4
    total_general = 0
    resumen_cc = {}  # para usarlo después en la hoja de detalle

    for cc in CentroCosto.objects.all():
        if cc_id and str(cc.id) != cc_id:
            continue

        ordenes_cc = qs_ordenes.filter(centro_costo=cc)
        detalles = DetalleOrden.objects.filter(orden__in=ordenes_cc)

        total_cc = 0
        for det in detalles:
            total_cc += float(det.cantidad * det.producto.precio)

        if total_cc == 0:
            continue

        resumen_cc[cc.id] = total_cc
        total_general += total_cc

        c1 = ws_resumen.cell(row=row_res, column=1, value=str(cc))
        c1.border = thin_border
        c2 = ws_resumen.cell(row=row_res, column=2, value=total_cc)
        c2.style = "moneda"
        c2.border = thin_border

        row_res += 1

    # Total general en resumen
    if row_res > 4:
        ws_resumen.cell(row=row_res + 1, column=1, value="Total general:").font = Font(bold=True)
        c_tot = ws_resumen.cell(row=row_res + 1, column=2, value=total_general)
        c_tot.style = "moneda"
        c_tot.font = Font(bold=True)

    # Ancho columnas resumen
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 20
    
    # Si hay datos en el resumen, crear gráfico de barras
    if row_res > 4:
        # Categorías: centros de costo (columna A, filas 4 hasta row_res-1)
        cats = Reference(ws_resumen, min_col=1, min_row=4, max_row=row_res-1)
        # Datos: totales (columna B, de 3 a row_res-1, incluyendo encabezado en 3)
        data = Reference(ws_resumen, min_col=2, min_row=3, max_row=row_res-1)

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Gasto por centro de costo"
        chart.y_axis.title = "Gasto"
        chart.x_axis.title = "Centro de costo"

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Mostrar valores sobre las barras
        chart.dataLabels = DataLabelList(showVal=True)

        # Opcional: ancho/alto para que no se vea aplastada
        chart.width = 24
        chart.height = 14

        ws_resumen.add_chart(chart, "D3")

    # ---------- HOJA 2: DETALLE ----------
    ws_detalle = wb.create_sheet(title="Detalle")

    # Título
    ws_detalle.merge_cells('A1:E1')
    ws_detalle['A1'] = "Detalle de gastos por centro de costo"
    ws_detalle['A1'].font = Font(size=14, bold=True)
    ws_detalle['A1'].alignment = Alignment(horizontal="center")

    # Filtros mostrados
    filtro_texto = []
    if fecha_inicio:
        filtro_texto.append(f"Desde: {fecha_inicio}")
    if fecha_fin:
        filtro_texto.append(f"Hasta: {fecha_fin}")
    if cc_id:
        cc_f = CentroCosto.objects.filter(id=cc_id).first()
        if cc_f:
            filtro_texto.append(f"Centro de costo: {cc_f}")
    ws_detalle['A2'] = " | ".join(filtro_texto) if filtro_texto else "Sin filtros"
    ws_detalle['A2'].alignment = Alignment(horizontal="left")

    row_det = 4  # empezamos un poco más abajo

    for cc in CentroCosto.objects.all():
        if cc_id and str(cc.id) != cc_id:
            continue

        ordenes_cc = qs_ordenes.filter(centro_costo=cc)
        detalles = DetalleOrden.objects.filter(orden__in=ordenes_cc)

        if not detalles.exists():
            continue

                # Título de bloque del centro de costo (alineado con la tabla, desde columna B)
        ws_detalle.merge_cells(start_row=row_det, start_column=2, end_row=row_det, end_column=5)
        cell_title = ws_detalle.cell(row=row_det, column=2, value=f"Centro de costo: {cc}")
        cell_title.font = Font(bold=True, color="003366")
        cell_title.alignment = Alignment(horizontal="left")
        row_det += 1

        # Encabezados del bloque (también desde columna B)
        headers_det = ['Producto', 'Cantidad', 'Precio', 'Gasto']
        for col, header in enumerate(headers_det, start=2):  # empezamos en columna B

            cell = ws_detalle.cell(row=row_det, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row_det += 1

        subtotal_cc = 0

        # Filas de detalle
        for det in detalles:
            costo = float(det.cantidad * det.producto.precio)
            subtotal_cc += costo

            ws_detalle.cell(row=row_det, column=2, value=det.producto.nombre).border = thin_border
            ws_detalle.cell(row=row_det, column=3, value=det.cantidad).border = thin_border

            c_precio = ws_detalle.cell(row=row_det, column=4, value=float(det.producto.precio))
            c_precio.style = "moneda"
            c_precio.border = thin_border

            c_gasto = ws_detalle.cell(row=row_det, column=5, value=costo)
            c_gasto.style = "moneda"
            c_gasto.border = thin_border

            row_det += 1

        # Subtotal del centro
        ws_detalle.merge_cells(start_row=row_det, start_column=2, end_row=row_det, end_column=4)
        cell_sub = ws_detalle.cell(row=row_det, column=2, value=f"Subtotal {cc}:")
        cell_sub.font = Font(bold=True)
        cell_sub.alignment = Alignment(horizontal="right")

        c_subtotal = ws_detalle.cell(row=row_det, column=5, value=subtotal_cc)
        c_subtotal.style = "moneda"
        c_subtotal.font = Font(bold=True)
        row_det += 2  # dejamos una fila en blanco antes del siguiente centro

    # Ancho columnas detalle
    for col in range(1, 6):
        ws_detalle.column_dimensions[get_column_letter(col)].width = 25


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_gastos.xlsx"'
    wb.save(response)
    return response




def lista_productos(request):
    productos = Producto.objects.all()
    return render(request, 'compras/lista_productos.html', {'productos': productos})

def editar_producto(request, id):
    producto = Producto.objects.get(id=id)
    if request.method == "POST":
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'compras/editar_producto.html', {'form': form})

def eliminar_producto(request, id):
    producto = Producto.objects.get(id=id)
    if request.method == "POST":
        producto.delete()
        return redirect('lista_productos')
    return render(request, 'compras/eliminar_producto.html', {'producto': producto})


def agregar_producto(request):
    if request.method == "POST":
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm()
    return render(request, 'compras/agregar_producto.html', {'form': form})

def agregar_orden_compra(request):
    if request.method == 'POST':
        orden_form = OrdenCompraForm(request.POST)
        detalle_form = DetalleOrdenForm(request.POST)
        if orden_form.is_valid() and detalle_form.is_valid():
            orden = orden_form.save()
            detalle = detalle_form.save(commit=False)
            detalle.orden = orden
            detalle.save()
            # Actualiza inventario del producto
            producto = Producto.objects.get(id=detalle.producto.id)
            producto.cantidad += detalle.cantidad
            producto.save()
            return redirect('lista_productos')
    else:
        orden_form = OrdenCompraForm()
        detalle_form = DetalleOrdenForm()
    return render(request, 'compras/agregar_orden.html', {
        'orden_form': orden_form,
        'detalle_form': detalle_form
    }) 
    
@login_required
@user_passes_test(es_compras)
def lista_ordenes(request):
    ordenes = OrdenCompra.objects.all().order_by('-fecha')
    proveedor = request.GET.get('proveedor', '').strip()
    estado = request.GET.get('estado', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    centro_costo = request.GET.get('centro_costo', '').strip()

    if proveedor:
        ordenes = ordenes.filter(proveedor__icontains=proveedor)

    if estado:
        ordenes = ordenes.filter(estado=estado)

    if fecha_desde:
        ordenes = ordenes.filter(fecha__gte=fecha_desde)

    if fecha_hasta:
        ordenes = ordenes.filter(fecha__lte=fecha_hasta)
        
    if centro_costo:
        # si es FK a CentroCosto y quieres filtrar por nombre
        ordenes = ordenes.filter(centro_costo__nombre__icontains=centro_costo)
        
    context = {
        'ordenes': ordenes,
        'f_proveedor': proveedor,
        'f_estado': estado,
        'f_fecha_desde': fecha_desde,
        'f_fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'compras/lista_ordenes.html', {'ordenes': ordenes})

def es_compras(user):
    return user.groups.filter(name='Compras').exists() or user.is_superuser

@login_required
@user_passes_test(es_compras)
def reporte_gastos_centrocosto(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    cc_id = request.GET.get('centro_costo')
    
    qs_ordenes = OrdenCompra.objects.all()
    if fecha_inicio:
        qs_ordenes = qs_ordenes.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs_ordenes = qs_ordenes.filter(fecha__lte=fecha_fin)
    if cc_id:
        qs_ordenes = qs_ordenes.filter(centro_costo__id=cc_id)
        
    cc_data = []
    for cc in CentroCosto.objects.all():
        if cc_id and str(cc.id) != cc_id:
            continue
        
        órdenes = qs_ordenes.filter(centro_costo=cc)
        total_gasto = 0
        detalles = DetalleOrden.objects.filter(orden__in=órdenes)
        desglose = []
        
        for det in detalles:
            costo = det.cantidad * det.producto.precio
            total_gasto += costo
            desglose.append({
                'producto': det.producto.nombre,
                'cantidad': det.cantidad,
                'precio': det.producto.precio,
                'gasto': costo,
            })
        cc_data.append({
            'centro_costo': cc,
            'total_gasto': total_gasto,
            'detalles': desglose
        })
    total_general = sum(cc['total_gasto'] for cc in cc_data)
    return render(request, 'compras/reporte_gastos_centrocosto.html', {
        'cc_data': cc_data,
        'total_general': total_general,
        'centros_costo': CentroCosto.objects.all()
    })


def detalle_orden(request, id):
    orden = get_object_or_404(OrdenCompra, id=id)
    detalles = DetalleOrden.objects.filter(orden=orden)
    subtotal = sum(d.cantidad * d.precio_unitario for d in detalles)
    iva = subtotal * Decimal ('0.16')
    total = subtotal - orden.descuento + iva - orden.anticipo
    return render(request, 'compras/orden_compra_detalle.html', {
        'orden': orden,
        'detalles': detalles,
        'subtotal': subtotal,
        'iva': iva,
        'total': total,
    })

def es_compras(user):
    return user.groups.filter(name='Compras').exists() or user.is_superuser

@login_required
@user_passes_test(es_compras)
def requisiciones_pendientes(request):
    # Aquí puedes filtrar por estatus, área, etc. según tu lógica.
    requisiciones = Requisicion.objects.filterl(estado="PENDIENTE").order_by('fecha')
    return render(request, 'compras/requisiciones_pendientes.html', {'requisiciones': requisiciones})


@login_required
@user_passes_test(es_compras)
def requisiciones_pendientes(request):
    requisiciones = Requisicion.objects.filter(estado="PENDIENTE").order_by('-fecha')
    return render(request, 'compras/requisiciones_pendientes.html', {'requisiciones': requisiciones})


@login_required
@user_passes_test(es_compras)
def aprobar_requisicion(request, id):
    requisicion = get_object_or_404(Requisicion, id=id)
    requisicion.estado = "APROBADA"
    requisicion.fecha_autorizacion = timezone.now().date()
    requisicion.save()
    messages.success(request, "Requisición aprobada.")
    return redirect('requisiciones_pendientes')


@login_required
@user_passes_test(es_compras)
def rechazar_requisicion(request, id):
    requisicion = get_object_or_404(Requisicion, id=id)
    requisicion.estado = "RECHAZADA"
    requisicion.fecha_autorizacion = timezone.now().date()
    requisicion.save()
    messages.warning(request, "Requisición rechazada.")
    return redirect('requisiciones_pendientes')

@login_required
def detalle_requisicion(request, id):
    requisicion = get_object_or_404(Requisicion, id=id)
    detalles = requisicion.detalles.all()  # related_name="detalles"
    return render(request, 'compras/detalle_requisicion.html', {
        'requisicion': requisicion,
        'detalles': detalles,
    })
    
    
@login_required
@user_passes_test(es_compras)
def recibir_orden(request, id):
    orden = get_object_or_404(OrdenCompra, id=id)

    if request.method == "POST":
        # Solo si aún no estaba recibida
        if orden.estado != "RECIBIDA":
            for det in orden.detalleorden_set.all():
                producto = det.producto
                existencia_antes = producto.cantidad
                producto.cantidad += det.cantidad
                producto.save()
                
                MovimientoInventario.objects.create(
                    producto=producto,
                    orden_compra=orden,
                    tipo="ENTRADA_OC",
                    cantidad=det.cantidad,
                    existencia_antes=existencia_antes,
                    existencia_despues=producto.cantidad,
                )
            orden.estado = "RECIBIDA"
            orden.save()
            messages.success(request, "Orden recibida y stock actualizado.")
        return redirect('lista_ordenes')

    return render(request, "compras/recibir_orden.html", {"orden": orden})

@login_required
def movimientos_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    movimientos = producto.movimientos.order_by('-fecha')
    return render(request, "compras/movimientos_producto.html", {
        "producto": producto,
        "movimientos": movimientos,
    })

@login_required
@user_passes_test(es_compras)
def reporte_compras_proveedor(request):
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    detalles = DetalleOrden.objects.select_related('orden', 'orden__centro_costo', 'orden__proveedor')

    if fecha_desde:
        detalles = detalles.filter(orden__fecha__gte=fecha_desde)
    if fecha_hasta:
        detalles = detalles.filter(orden__fecha__lte=fecha_hasta)

    importe_expr = ExpressionWrapper(
        F('cantidad') * F('precio_unitario'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    resumen = (
        detalles
        .values('orden__proveedor')
        .annotate(total_comprado=Sum(importe_expr))
        .order_by('orden__proveedor')
    )

    context = {
        'resumen': resumen,
        'f_fecha_desde': fecha_desde,
        'f_fecha_hasta': fecha_hasta,
    }
    return render(request, 'compras/reporte_compras_proveedor.html', context)
